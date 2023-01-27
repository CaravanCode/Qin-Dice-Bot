import discord
from discord.ext import commands
import random
import re
import os
from openpyxl import load_workbook


bot = commands.Bot(command_prefix="!", intents=discord.Intents.all())
TOKEN = "Your Bot Token"
emo_yin = "<:yin:1064122799091884082>"
emo_yang = "<:yang:1064122752866463775>"
emo_yy = "<:yin_yang:1064348918672015380>"
emo_yin_m = "<:yin_m:1065391100132605953>"
emo_yang_m = "<:yang_m:1065391125176799343>"


@bot.event
async def on_ready():
    print("Qin Dice Bot is ready!")


@bot.command()
async def weapon(ctx, *args):
    weapon_result = load_weapon()
    if not args:
        separator = ', '
        weapons = separator.join(weapon_result[0])
        weapon_list = weapons.split(separator)
        weapon_list.sort()
        result = '\n'.join(weapon_list)
        await ctx.send(f"{ctx.author.mention}"+"\n"+result)
    else:
        filtered_weapons = [x for x in weapon_result[0]
                            if args[0].lower() == x.lower()]
        if filtered_weapons:
            result = choosen_weapon(filtered_weapons)
            weapon_name = result[0]
            weapon_name = weapon_name.title()
            weapon_skill = str(result[3])
            weapon_skill = weapon_skill.title()
            embed = discord.Embed(
                title=weapon_name, description=result[5], color=0xFF5733)
            embed.add_field(name="Damage", value=result[1], inline=True)
            embed.add_field(name="Resilience", value=result[2], inline=True)
            embed.add_field(name="Skill", value=weapon_skill, inline=True)
            await ctx.send(f"{ctx.author.mention}\n", embed=embed)
        else:
            await ctx.send("Weapon not found.")


@bot.command()
async def s(ctx, *args):
    try:
        args = " ".join(args)
        input_list = re.split(r'[+ ]', args)
        num_dice, num_sides = input_list[0].split("d")
        modifiers = [int(i)
                     for i in input_list[1:] if i.lstrip("+-").isdigit()]
        desc = [i for i in input_list[1:] if (
            not i.lstrip("+-").isdigit()) and i != '']
        result = skill_check(int(num_dice), int(num_sides), modifiers, desc)
        await ctx.send(f"{ctx.author.mention}"+"\n"+result)
    except ValueError:
        await ctx.send("Error")


@bot.command()
async def a(ctx, *args):
    try:
        args = " ".join(args)
        input_list = re.split(r'[+ ]', args)
        num_dice, num_sides = input_list[0].split("d")
        modifiers = [int(i)
                     for i in input_list[1:] if i.lstrip("+-").isdigit()]
        desc = [i for i in input_list[1:] if (
            not i.lstrip("+-").isdigit()) and i != '']
        weapon_result = choosen_weapon(desc)
        result = attack_check(int(num_dice), int(num_sides),
                              modifiers, weapon_result[0], int(weapon_result[1]))
        await ctx.send(f"{ctx.author.mention}"+"\n"+result)
    except ValueError as e:
        await ctx.send("Invalid input. Please enter in the format '!a 2d10 (metal) (martial skill) (registered weapon)'")


def roll_dice2(num_dice, num_sides):
    rolls = []
    for i in range(num_dice):
        rolls.append(random.randint(0, num_sides-1))
    yin = rolls[0]
    yang = rolls[1]
    return yin, yang


def dice_checker(yin, yang):
    if yin == yang:
        if yin == 0:
            result = "Crit Fail"
            dice_result = 0
            return result, yin
        else:
            result = "Crit Success"
            dice_result = yin
            return result, dice_result
    else:
        if yin > yang:
            result = 0
            dice_result = yin - yang
            return result, dice_result
        else:
            result = 1
            dice_result = yang - yin
            return result, dice_result


def skill_check(num_dice, num_sides, modifiers, description):
    desc = " ".join(description)
    yin, yang = roll_dice2(num_dice, num_sides)
    checker, dice_result = dice_checker(yin, yang)
    sum_mod = sum(modifiers)
    total = dice_result + sum_mod
    if checker == "Crit Fail":
        result = f"**{desc}: \n{yin} {emo_yin} {yang} {emo_yang} Critical Fail!\nResult: 0 {emo_yy}\nChi Lost: 5**"
        return result
    elif checker == "Crit Success":
        if sum_mod < 0:
            result = f"**{desc}: \n{yin} {emo_yin} {yang} {emo_yang} Critical Success!\nResult: {yin} {emo_yy} {sum_mod} = {total}\nChi Gainned: {yin}**"
            return result
        else:
            result = f"**{desc}: \n{yin} {emo_yin} {yang} {emo_yang} Critical Success!\nResult: {yin} {emo_yy} + {sum_mod} = {total}\nChi Gainned: {yin}**"
            return result
    else:
        if checker == 0:
            result = f"**{desc}: **\n{yin} {emo_yin} {yang} {emo_yang}\n**Result: ** {dice_result} {emo_yin_m} + {sum_mod} = {total}"
            return result
        else:
            result = f"**{desc}: **\n{yin} {emo_yin} {yang} {emo_yang}\n**Result: ** {dice_result} {emo_yang_m} + {sum_mod} = {total}"
            return result


def choosen_weapon(weapon):
    weapon_names, weapon_descs, weapon_damages, weapon_resiliences, weapon_skills, weapon_prices = load_weapon()
    weapon_name = weapon[0]
    if weapon_name in weapon_names:
        # Do something with the weapon
        index = weapon_names.index(weapon_name)
        alt_name = weapon_name.split("_")
        if len(alt_name) == 1:
            # case where input is "bang"
            name_join = " ".join(alt_name)
            weapon_name = name_join
            return weapon_name, weapon_damages[index], weapon_resiliences[index], weapon_skills[index], weapon_prices[index], weapon_descs[index]
        else:
            # case where input is "bang_xiao"
            alt_name = alt_name[::-1]
            name_join = " ".join(alt_name)
            weapon_name = name_join
            return weapon_name, weapon_damages[index], weapon_resiliences[index], weapon_skills[index], weapon_prices[index], weapon_descs[index]
    else:
        print("Invalid weapon")
        return None, None


def damage_check(num_dice, num_sides, modifiers, weapon_name, weapon_damage):
    damage_yin, damage_yang = roll_dice2(num_dice, num_sides)
    damage_checker, damage_dice_result = dice_checker(damage_yin, damage_yang)
    print(damage_checker)
    if (damage_checker == ("Crit Fail")):
        total_damage = modifiers + weapon_damage + damage_dice_result
        return f"**Damage:**\n{damage_yin} {emo_yin} {damage_yang} {emo_yang}\n**Result: **{damage_dice_result} {emo_yy} + {modifiers}(Metal) + {weapon_damage}({weapon_name.title()}) = {total_damage}"
    elif (damage_checker == ("Crit Success")):
        total_damage = modifiers + weapon_damage + damage_dice_result
        return f"**Damage:**\n{damage_yin} {emo_yin} {damage_yang} {emo_yang}\n**Result: **{damage_dice_result} {emo_yy} + {modifiers}(Metal) + {weapon_damage}({weapon_name.title()}) = {total_damage}"
    else:
        if (damage_checker == 1):
            total_damage = modifiers + weapon_damage + damage_dice_result
            return f"**Damage:**\n{damage_yin} {emo_yin} {damage_yang} {emo_yang}\n**Result: **{damage_dice_result} {emo_yang_m} + {modifiers}(Metal) + {weapon_damage}({weapon_name.title()}) = {total_damage}"
        else:
            total_damage = modifiers + weapon_damage
            return f"**Damage:**\n{damage_yin} {emo_yin} {damage_yang} {emo_yang}\n**Result: **{modifiers}(Metal) + {weapon_damage}({weapon_name.title()}) = {total_damage}"


def attack_check(num_dice, num_sides, modifiers, weapon_name, weapon_damage):
    attack_yin, attack_yang = roll_dice2(num_dice, num_sides)
    attack_checker, attack_dice_result = dice_checker(attack_yin, attack_yang)
    attack_mod = sum(modifiers)
    damage_mod = modifiers[0]
    total = attack_dice_result + attack_mod
    damage_result = damage_check(
        int(num_dice), int(num_sides), int(damage_mod), weapon_name, int(weapon_damage))

    if attack_checker == "Crit Fail":
        result = f"**Attack with {weapon_name.title()}! \n{attack_yin} {emo_yin} {attack_yang} {emo_yang} Critical Fail!\nResult: 0 {emo_yy}**"
        return result
    elif attack_checker == "Crit Success":
        if attack_mod < 0:
            result = f"**Attack with {weapon_name.title()}! \n{attack_yin} {emo_yin} {attack_yang} {emo_yang} Critical Success!\nResult: {attack_yin} {emo_yy} {attack_mod} = {total}**\n\n{damage_result}"
            return result
        else:
            result = f"**Attack with {weapon_name.title()}! \n{attack_yin} {emo_yin} {attack_yang} {emo_yang} Critical Success!\nResult: {attack_yin} {emo_yy} + {attack_mod} = {total}**\n\n{damage_result}"
            return result
    else:
        if attack_checker == 0:
            result = f"Attack with {weapon_name.title()}! \n{attack_yin} {emo_yin} {attack_yang} {emo_yang}\n**Result: ** {attack_dice_result} {emo_yin_m} + {attack_mod} = {total}\n\n{damage_result}"
            return result
        else:
            result = f"Attack with {weapon_name.title()}! \n{attack_yin} {emo_yin} {attack_yang} {emo_yang}\n**Result: ** {attack_dice_result} {emo_yang_m} + {attack_mod} = {total}\n\n{damage_result}"
            return result


def load_weapon():
    file = os.path.join(os.path.dirname(__file__), 'weapons.xlsx')
    workbook = load_workbook(file)

    sheet_name = 'weapons'

    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        weapon_name = []
        weapon_desc = []
        weapon_damage = []
        weapon_resilience = []
        weapon_skill = []
        weapon_price = []

        # Print the data starting from the B3
        row_index = 0
        for row in sheet.iter_rows(values_only=True):
            if row_index >= 2:
                # print(row)
                weapon_name.append(row[0])
                weapon_desc.append(row[1])
                weapon_damage.append(row[2])
                weapon_resilience.append(row[3])
                weapon_skill.append(row[4])
                weapon_price.append(row[5])
            row_index += 1
        return weapon_name, weapon_desc, weapon_damage, weapon_resilience, weapon_skill, weapon_price
    else:
        return None, None, None, None, None, None


bot.run(TOKEN)
